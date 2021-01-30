
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb=Workbook()
# dest_filename="empty_book1.xlsx"
# ws1=wb.active
# ws1.title="range names"
# for row in range(1,40):
#     ws1.append(range(600))
# ws2=wb.create_sheet(title="1")
# ws2['F1']=3.34
# ws3=wb.create_sheet(title="2")
# for row in range(1,21):
#     for col in range(1,21):
#         ws3.cell(row=row,column=col,value="{0}".format(get_column_letter))
# wb.save(filename=dest_filename)
lw1=load_workbook(filename="empty_book1.xlsx")
lw1_sheet=lw1["range names"]
for i in range(1,20):
    print(lw1_sheet["A{}".format(i)].value)

